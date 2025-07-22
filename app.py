from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
import pandas as pd
import os
import uuid
import tempfile
import shutil
from datetime import datetime, timedelta
from openpyxl.styles import Alignment, numbers
from openpyxl import Workbook
from werkzeug.utils import secure_filename
import threading
import time

app = Flask(__name__)
app.secret_key = 'your-secret-key-change-this-in-production'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Configuration
UPLOAD_FOLDER = 'uploads'
OUTPUT_FOLDER = 'output'
ALLOWED_EXTENSIONS = {'txt'}

# Create necessary directories
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)
os.makedirs('static', exist_ok=True)  # For logo and static assets

# Store processing history and file cache in memory
processing_history = {}
file_cache = {}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def cleanup_old_files():
    """Clean up files older than 1 hour"""
    while True:
        try:
            current_time = datetime.now()
            for folder in [UPLOAD_FOLDER, OUTPUT_FOLDER]:
                for filename in os.listdir(folder):
                    file_path = os.path.join(folder, filename)
                    if os.path.isfile(file_path):
                        file_time = datetime.fromtimestamp(os.path.getctime(file_path))
                        if current_time - file_time > timedelta(hours=1):
                            os.remove(file_path)
            time.sleep(3600)
        except Exception as e:
            print(f"Cleanup error: {e}")
            time.sleep(3600)

cleanup_thread = threading.Thread(target=cleanup_old_files, daemon=True)
cleanup_thread.start()

def parse_candidate_line(line):
    roll_no = line[:8].strip()
    remaining = line[8:].lstrip()
    if not remaining:
        return roll_no, '', '', [], '', ''
    
    gender = remaining[0]
    remaining = remaining[1:].lstrip()
    parts = remaining.split()
    name_parts = []
    subjects = []
    result = ''
    comp_sub = []
    found_subject = False
    
    i = 0
    while i < len(parts):
        token = parts[i]
        if len(token) == 3 and token.isdigit():
            found_subject = True
            subjects.append(token)
        elif not found_subject:
            name_parts.append(token)
        else:
            if token == 'ESSENTIAL' and i + 1 < len(parts) and parts[i+1] == 'REPEAT':
                result = 'ESSENTIAL REPEAT'
                i += 1
                comp_sub = parts[i+1:]
                break
            elif token in ['PASS', 'COMP', 'UFM', 'ABST', 'REPEAT']:
                result = token
                comp_sub = parts[i+1:]
                break
        i += 1
    
    name = ' '.join(name_parts)
    return roll_no, gender, name, subjects, result, ' '.join(comp_sub)

def parse_marks_line(line):
    tokens = line.split()
    marks_and_grades = []
    
    i = 0
    while i < len(tokens):
        if tokens[i].replace('-', '').isdigit() or tokens[i] == '':
            if i + 1 < len(tokens):
                mark = tokens[i].lstrip('0') or '0'
                marks_and_grades.append((mark, tokens[i+1]))
                i += 2
            else:
                mark = tokens[i].lstrip('0') or '0'
                marks_and_grades.append((mark, ''))
                i += 1
        else:
            i += 1
    
    return marks_and_grades

# Enhanced function to parse and cache data
def parse_and_cache_file(input_file):
    """Parse file and cache the structured data for faster filtering"""
    cache_key = f"{input_file}_{os.path.getmtime(input_file)}"
    
    if cache_key in file_cache:
        return file_cache[cache_key]
    
    with open(input_file, 'r', encoding='utf-8') as f:
        lines = [line.rstrip('\n').rstrip('\r') for line in f.readlines()]
    
    all_subjects = set()
    candidates = []
    
    # First pass: collect all subjects
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        if line and line[:8].strip().isdigit():
            _, _, _, subjects, _, _ = parse_candidate_line(line)
            all_subjects.update(subjects)
        i += 1
    
    subject_codes = sorted(all_subjects)
    
    # Second pass: parse all candidates
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        if line and line[:8].strip().isdigit():
            roll_no, gender, name, subjects, result, comp_sub = parse_candidate_line(line)
            
            i += 1
            marks_line = lines[i].strip() if i < len(lines) else ''
            marks_and_grades = parse_marks_line(marks_line)
            
            candidate_data = {
                'Roll No': roll_no,
                'Gender': gender,
                'Name': name
            }
            
            for code in subject_codes:
                candidate_data[f"{code}_Marks"] = ''
                candidate_data[f"{code}_Grade"] = ''
            
            if result not in ['UFM', 'ABST']:
                for subj, (mark, grade) in zip(subjects, marks_and_grades):
                    if subj in subject_codes:
                        candidate_data[f"{subj}_Marks"] = int(mark) if mark.isdigit() else mark
                        candidate_data[f"{subj}_Grade"] = grade
            
            candidate_data['Result'] = result
            candidate_data['Comp Sub'] = comp_sub
            candidates.append(candidate_data)
        
        i += 1
    
    columns = ['Roll No', 'Gender', 'Name']
    for code in subject_codes:
        columns.append(f"{code}_Marks")
        columns.append(f"{code}_Grade")
    columns.extend(['Result', 'Comp Sub'])
    
    cached_data = {
        'candidates': candidates,
        'columns': columns,
        'subject_codes': subject_codes
    }
    
    file_cache[cache_key] = cached_data
    return cached_data


# Add this function after parse_and_cache_file function
def remove_empty_columns_from_df(df):
    """Remove columns where all values are blank (empty strings, NaNs, or whitespace)."""
    try:
        # Make a copy to avoid modifying the original
        df_clean = df.copy()
        
        # Keep essential columns that should never be removed
        essential_columns = ['Roll No', 'Gender', 'Name', 'Result', 'Comp Sub']
        
        # Identify columns to potentially remove (subject columns)
        columns_to_check = [col for col in df_clean.columns if col not in essential_columns]
        
        # Drop only subject columns where all values are effectively empty
        columns_to_drop = []
        for col in columns_to_check:
            # Check if all values are empty (None, NaN, empty string, or whitespace)
            is_empty = df_clean[col].apply(lambda x: 
                x is None or 
                pd.isna(x) or 
                str(x).strip() == '' or 
                str(x).strip().lower() == 'nan'
            )
            
            if is_empty.all():
                columns_to_drop.append(col)
        
        # Drop the empty columns
        df_clean = df_clean.drop(columns=columns_to_drop)
        
        print(f"Removed {len(columns_to_drop)} empty columns: {columns_to_drop}")  # Debug info
        
        return df_clean
    
    except Exception as e:
        print(f"Error in remove_empty_columns_from_df: {e}")
        # Return original dataframe if there's an error
        return df

# Fast filtering function for single filter
def create_filtered_excel(cached_data, filter_roll_numbers, sheet_name="Filtered"):
    """Create Excel file with filtered data using cached parsed data"""
    if not filter_roll_numbers:
        return None, 0
    
    # Filter candidates
    filtered_candidates = [
        candidate for candidate in cached_data['candidates']
        if candidate['Roll No'] in filter_roll_numbers
    ]
    
    if not filtered_candidates:
        return None, 0
    
    # Create DataFrame and remove empty columns
    df_filtered = pd.DataFrame(filtered_candidates, columns=cached_data['columns'])
    df_filtered = remove_empty_columns_from_df(df_filtered)
    
    # Create Excel file
    output_file = os.path.join(OUTPUT_FOLDER, f"filtered_{uuid.uuid4().hex[:8]}.xlsx")
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_filtered.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Apply formatting
        workbook = writer.book
        ws = workbook[sheet_name]
        
        for col in ws.iter_cols():
            column_letter = col[0].column_letter
            column_name = col[0].value
            
            if isinstance(column_name, str) and column_name.endswith('_Marks'):
                for cell in col:
                    if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.isdigit()):
                        cell.number_format = numbers.FORMAT_NUMBER
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                for cell in col:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
    
    return output_file, len(filtered_candidates)

 
# Multi-filter function for creating multiple sheets
def create_multi_filtered_excel(cached_data, filter_sets):
    """Create Excel file with multiple filtered sheets"""
    if not filter_sets:
        return None, 0
    
    # Create Excel file
    output_file = os.path.join(OUTPUT_FOLDER, f"multi_filter_{uuid.uuid4().hex[:8]}.xlsx")
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # First add all students sheet (keep all columns for reference)
        df_all = pd.DataFrame(cached_data['candidates'], columns=cached_data['columns'])
        df_all.to_excel(writer, sheet_name='All Students', index=False)
        
        sheets_created = 0
        total_filtered = 0
        
        # Create filtered sheets
        for idx, roll_block in enumerate(filter_sets, start=1):
            # Parse roll numbers from block
            rolls = [r.strip() for r in roll_block.replace(',', '\n').split('\n') if r.strip()]
            if not rolls:
                continue
                
            # Filter candidates
            filtered_candidates = [
                candidate for candidate in cached_data['candidates']
                if candidate['Roll No'] in rolls
            ]
            
            if not filtered_candidates:
                continue
                
            # Create DataFrame and remove empty columns for this filtered set
            df_filtered = pd.DataFrame(filtered_candidates, columns=cached_data['columns'])
            df_filtered = remove_empty_columns_from_df(df_filtered)
            
            # Create sheet
            sheet_name = f"Filter_{idx}"
            df_filtered.to_excel(writer, sheet_name=sheet_name, index=False)
            
            sheets_created += 1
            total_filtered += len(filtered_candidates)
        
        # Apply formatting to all sheets
        workbook = writer.book
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            for col in ws.iter_cols():
                column_letter = col[0].column_letter
                column_name = col[0].value
                
                if isinstance(column_name, str) and column_name.endswith('_Marks'):
                    for cell in col:
                        if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.isdigit()):
                            cell.number_format = numbers.FORMAT_NUMBER
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:
                    for cell in col:
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = adjusted_width
    
    return output_file, sheets_created, total_filtered

# Original text_to_excel function for full processing
def text_to_excel(input_file, filter_roll_numbers=None):
    if filter_roll_numbers is None:
        filter_roll_numbers = []
    
    base_name = os.path.splitext(os.path.basename(input_file))[0]
    output_file = os.path.join(OUTPUT_FOLDER, f"{base_name}_{uuid.uuid4().hex[:8]}.xlsx")
    
    with open(input_file, 'r', encoding='utf-8') as f:
        lines = [line.rstrip('\n').rstrip('\r') for line in f.readlines()]
    
    all_subjects = set()
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        if line and line[:8].strip().isdigit():
            _, _, _, subjects, _, _ = parse_candidate_line(line)
            all_subjects.update(subjects)
        i += 1
    
    subject_codes = sorted(all_subjects)
    candidates = []
    filtered_candidates = []
    stats = {
        'TOTAL': 0,
        'PASS': 0,
        'COMP': 0,
        'ESSENTIAL REPEAT': 0,
        'ABST': 0,
        'UFM': 0,
        'OTHER': 0
    }
    
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        if line and line[:8].strip().isdigit():
            stats['TOTAL'] += 1
            roll_no, gender, name, subjects, result, comp_sub = parse_candidate_line(line)
            
            key = result if result in stats else 'OTHER'
            if key in stats:
                stats[key] += 1
            
            i += 1
            marks_line = lines[i].strip() if i < len(lines) else ''
            marks_and_grades = parse_marks_line(marks_line)
            
            candidate_data = {
                'Roll No': roll_no,
                'Gender': gender,
                'Name': name
            }
            
            for code in subject_codes:
                candidate_data[f"{code}_Marks"] = ''
                candidate_data[f"{code}_Grade"] = ''
            
            if result not in ['UFM', 'ABST']:
                for subj, (mark, grade) in zip(subjects, marks_and_grades):
                    if subj in subject_codes:
                        candidate_data[f"{subj}_Marks"] = int(mark) if mark.isdigit() else mark
                        candidate_data[f"{subj}_Grade"] = grade
            
            candidate_data['Result'] = result
            candidate_data['Comp Sub'] = comp_sub
            candidates.append(candidate_data)
            
            if roll_no in filter_roll_numbers:
                filtered_candidates.append(candidate_data)
        
        i += 1
    
    columns = ['Roll No', 'Gender', 'Name']
    for code in subject_codes:
        columns.append(f"{code}_Marks")
        columns.append(f"{code}_Grade")
    columns.extend(['Result', 'Comp Sub'])
    
    df = pd.DataFrame(candidates, columns=columns)
    df_filtered = pd.DataFrame(filtered_candidates, columns=columns)


    
    
    # Remove empty columns from filtered data only (keep all columns in "All Students")
    if len(filtered_candidates) > 0:
        df_filtered = remove_empty_columns_from_df(df_filtered)
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='All Students', index=False)
        if len(filtered_candidates) > 0:
            df_filtered.to_excel(writer, sheet_name='Filtered Students', index=False)
        
        workbook = writer.book
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            for col in ws.iter_cols():
                column_letter = col[0].column_letter
                column_name = col[0].value
                if isinstance(column_name, str) and column_name.endswith('_Marks'):
                    for cell in col:
                        if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.isdigit()):
                            cell.number_format = numbers.FORMAT_NUMBER
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:
                    for cell in col:
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = adjusted_width
    
    return output_file, stats, len(filtered_candidates)

# Routes
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file selected'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'error': 'Invalid file type. Only .txt files are allowed'}), 400
        
        filename = secure_filename(file.filename)
        unique_filename = f"{uuid.uuid4().hex[:8]}_{filename}"
        filepath = os.path.join(UPLOAD_FOLDER, unique_filename)
        file.save(filepath)
        
        # Parse and cache the file immediately for fast filtering
        cached_data = parse_and_cache_file(filepath)
        
        session['uploaded_file'] = filepath
        session['original_filename'] = filename
        
        return jsonify({
            'message': 'File uploaded and parsed successfully', 
            'filename': filename,
            'total_students': len(cached_data['candidates'])
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Dynamic filter endpoint
@app.route('/filter_dynamic', methods=['POST'])
def filter_dynamic():
    """Handle dynamic filtering with real-time Excel creation"""
    try:
        if 'uploaded_file' not in session:
            return jsonify({'error': 'No file uploaded'}), 400
        
        filepath = session['uploaded_file']
        if not os.path.exists(filepath):
            return jsonify({'error': 'Uploaded file not found'}), 400
        
        # Get roll numbers from request
        data = request.get_json()
        filter_text = data.get('roll_numbers', '')
        
        if not filter_text.strip():
            return jsonify({'error': 'No roll numbers provided'}), 400
        
        # Parse roll numbers
        rolls = [roll.strip() for roll in filter_text.replace(',', '\n').split('\n')]
        filter_roll_numbers = [roll for roll in rolls if roll]
        
        # Get cached data
        cached_data = parse_and_cache_file(filepath)
        
        # Create filtered Excel
        output_file, filtered_count = create_filtered_excel(
            cached_data, 
            filter_roll_numbers,
            f"Filtered_{len(filter_roll_numbers)}_Rolls"
        )
        
        if not output_file:
            return jsonify({'error': 'No matching students found'}), 404
        
        # Store result
        process_id = uuid.uuid4().hex[:8]
        processing_history[process_id] = {
            'timestamp': datetime.now().isoformat(),
            'original_filename': session.get('original_filename', 'unknown'),
            'output_file': output_file,
            'filtered_count': filtered_count,
            'filter_roll_numbers': filter_roll_numbers,
            'type': 'dynamic_filter'
        }
        
        return jsonify({
            'process_id': process_id,
            'filtered_count': filtered_count,
            'download_url': url_for('download_file', process_id=process_id),
            'message': f'Found {filtered_count} matching students'
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Multi-filter endpoint
@app.route('/filter_multi', methods=['POST'])
def filter_multi():
    """Handle multiple roll number sets and create multiple sheets"""
    try:
        if 'uploaded_file' not in session:
            return jsonify({'error': 'No file uploaded'}), 400
        
        filepath = session['uploaded_file']
        if not os.path.exists(filepath):
            return jsonify({'error': 'Uploaded file not found'}), 400
        
        # Parse payload
        data = request.get_json()
        filter_sets = data.get('sets', [])
        
        if not filter_sets:
            return jsonify({'error': 'No roll-number sets provided'}), 400
        
        # Get cached data
        cached_data = parse_and_cache_file(filepath)
        
        # Create multi-filtered Excel
        output_file, sheets_created, total_filtered = create_multi_filtered_excel(cached_data, filter_sets)
        
        if not output_file:
            return jsonify({'error': 'No matching students found in any set'}), 404
        
        # Store result
        process_id = uuid.uuid4().hex[:8]
        processing_history[process_id] = {
            'timestamp': datetime.now().isoformat(),
            'original_filename': session.get('original_filename', 'unknown'),
            'output_file': output_file,
            'sheets_created': sheets_created,
            'total_filtered': total_filtered,
            'filter_sets': filter_sets,
            'type': 'multi_filter'
        }
        
        return jsonify({
            'process_id': process_id,
            'sheets_created': sheets_created,
            'total_filtered': total_filtered,
            'download_url': url_for('download_file', process_id=process_id),
            'message': f'Created {sheets_created} filtered sheets with {total_filtered} total students'
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/process', methods=['POST'])
def process_file():
    try:
        if 'uploaded_file' not in session:
            return jsonify({'error': 'No file uploaded'}), 400
        
        filepath = session['uploaded_file']
        if not os.path.exists(filepath):
            return jsonify({'error': 'Uploaded file not found'}), 400
        
        filter_text = request.json.get('filter_roll_numbers', '')
        filter_roll_numbers = []
        if filter_text.strip():
            rolls = [roll.strip() for roll in filter_text.replace(',', '\n').split('\n')]
            filter_roll_numbers = [roll for roll in rolls if roll]
        
        output_file, stats, filtered_count = text_to_excel(filepath, filter_roll_numbers)
        
        process_id = uuid.uuid4().hex[:8]
        processing_history[process_id] = {
            'timestamp': datetime.now().isoformat(),
            'original_filename': session.get('original_filename', 'unknown'),
            'output_file': output_file,
            'stats': stats,
            'filtered_count': filtered_count,
            'filter_roll_numbers': filter_roll_numbers,
            'type': 'full_process'
        }
        
        session['last_process_id'] = process_id
        
        return jsonify({
            'process_id': process_id,
            'stats': stats,
            'filtered_count': filtered_count,
            'message': 'File processed successfully'
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/download/<process_id>')
def download_file(process_id):
    try:
        if process_id not in processing_history:
            return jsonify({'error': 'Process not found'}), 404
        
        process_info = processing_history[process_id]
        output_file = process_info['output_file']
        
        if not os.path.exists(output_file):
            return jsonify({'error': 'Output file not found'}), 404
        
        # Fix the download name construction
        if process_info.get('type') == 'dynamic_filter':
            download_name = f"filtered_{process_info['filtered_count']}_students.xlsx"
        elif process_info.get('type') == 'multi_filter':
            download_name = f"multi_filter_{process_info['sheets_created']}_sheets.xlsx"
        else:  # full_process
            # Ensure .xlsx extension for full processing
            base_name = os.path.splitext(process_info['original_filename'])[0]
            download_name = f"processed_{base_name}_complete.xlsx"
        
        return send_file(
            output_file, 
            as_attachment=True, 
            download_name=download_name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/history')
def get_history():
    try:
        recent_history = dict(list(processing_history.items())[-10:])
        return jsonify(recent_history), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/preview/<process_id>')
def preview_data(process_id):
    try:
        if process_id not in processing_history:
            return jsonify({'error': 'Process not found'}), 404
        
        process_info = processing_history[process_id]
        output_file = process_info['output_file']
        
        if not os.path.exists(output_file):
            return jsonify({'error': 'Output file not found'}), 404
        
        # Read first few rows
        df = pd.read_excel(output_file)
        preview_data = {
            'students': df.head(10).to_dict('records'),
            'total_students': len(df),
            'columns': df.columns.tolist()
        }
        
        return jsonify(preview_data), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Add these new routes after your existing routes

@app.route('/delete_upload', methods=['POST'])
def delete_upload():
    """Delete the currently uploaded file and clear session data"""
    try:
        if 'uploaded_file' not in session:
            return jsonify({'error': 'No file to delete'}), 400
        
        filepath = session['uploaded_file']
        
        # Delete the physical file if it exists
        if os.path.exists(filepath):
            os.remove(filepath)
        
        # Clear file cache entry
        cache_key = f"{filepath}_{os.path.getmtime(filepath) if os.path.exists(filepath) else 0}"
        if cache_key in file_cache:
            del file_cache[cache_key]
        
        # Clear session data
        session.pop('uploaded_file', None)
        session.pop('original_filename', None)
        session.pop('last_process_id', None)
        
        return jsonify({
            'message': 'Uploaded file deleted successfully',
            'success': True
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/clear_history', methods=['POST'])
def clear_history():
    """Clear all download history and associated files"""
    try:
        # Delete all output files from processing history
        deleted_files = 0
        for process_id, info in processing_history.items():
            output_file = info.get('output_file')
            if output_file and os.path.exists(output_file):
                os.remove(output_file)
                deleted_files += 1
        
        # Clear processing history
        processing_history.clear()
        
        return jsonify({
            'message': f'History cleared successfully. Deleted {deleted_files} files.',
            'deleted_files': deleted_files,
            'success': True
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/delete_history_item/<process_id>', methods=['POST'])
def delete_history_item(process_id):
    """Delete a specific history item and its associated file"""
    try:
        if process_id not in processing_history:
            return jsonify({'error': 'Process not found'}), 404
        
        # Delete the output file
        process_info = processing_history[process_id]
        output_file = process_info.get('output_file')
        
        if output_file and os.path.exists(output_file):
            os.remove(output_file)
        
        # Remove from history
        del processing_history[process_id]
        
        return jsonify({
            'message': 'History item deleted successfully',
            'success': True
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500







if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)
